import openpyxl

book = openpyxl.load_workbook("C:\\Users\\julian.patino\\Documents\\hola\\datosPyton.xlsx")
sheet = book.active
Dict = {}
cell = sheet.cell(row=1, column=2)
print(cell.value)
cell2 = sheet.cell(row=2, column=2).value = "Julian"
print(cell2)
print("Maximo numero de filas es", sheet.max_row)
print("Maximo numero de columnas es", sheet.max_column)
print(sheet['A4'].value)
print("_________________________________________________________________")
for element in range(1, sheet.max_row + 1):
    if sheet.cell(row=element, column=1).value == "Testcase2":
        for j in range(1, sheet.max_column + 1):
            # print(sheet.cell(row=element, column=j).value)
            Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=element, column=j).value
            print(Dict)
