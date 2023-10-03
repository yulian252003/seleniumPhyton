import openpyxl


class HomePageData:
    test_HomePage_data = [{"firstname": "Rahul", "lastname": "Shetty", "gender": "Male"},
                          {"firstname": "Andrea", "lastname": "Rios", "gender": "Female"}]

    @staticmethod
    def getTestData(test_case_name):
        book = openpyxl.load_workbook("C:\\Users\\julian.patino\\Documents\\hola\\datosPyton.xlsx")
        sheet = book.active
        dict_t = {}
        for element in range(1, sheet.max_row + 1):
            if sheet.cell(row=element, column=1).value == test_case_name:
                for j in range(2, sheet.max_column + 1):
                    dict_t[sheet.cell(row=1, column=j).value] = sheet.cell(row=element, column=j).value

        return [dict_t]
